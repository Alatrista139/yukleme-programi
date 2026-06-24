import sys
import math
import pandas as pd
from openpyxl import load_workbook
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QLabel, QComboBox, QSpinBox,
                             QPushButton, QTableWidget, QTableWidgetItem,
                             QMessageBox, QHeaderView, QMenu, QInputDialog,
                             QLineEdit, QDialog, QRadioButton,
                             QButtonGroup, QGroupBox)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QColor, QFont

CBM_EXCEL    = "CBM.xlsx"
TANIM_EXCEL  = "set_tanimlari.xlsx"
PAKET_KATSAYI = 1.02

AYAK_KOLI_HACIM        = round(39 * 75 * 30 / 1_000_000, 6)
AYAK_KOLI_KAPASITE     = 144
AYAK_KUCUK_GENISLIKLER = {70, 80, 90, 100, 110, 120}

ARAC_BOYUTLARI = {
    "40HC Konteyner": {"en": 235, "boy": 1200, "yukseklik": 270},
    "Tir":            {"en": 245, "boy": 1360, "yukseklik": 290},
}


def cbm_tablosundan_boyut_cek(sayfa_adi, model_adi, olcu_sayisi):
    try:
        wb = load_workbook(CBM_EXCEL, read_only=True, data_only=True)
        gercek_sayfa = next((s for s in wb.sheetnames if s.upper() == sayfa_adi.upper()), None)
        if gercek_sayfa is None:
            return None
        ws   = wb[gercek_sayfa]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None
        olcu_row     = rows[0]
        model_aranan = str(model_adi).replace(" ", "").upper()
        model_satiri = None
        for r in rows[2:]:
            if r[0] and str(r[0]).replace(" ", "").upper() == model_aranan:
                model_satiri = r
                break
        if model_satiri is None:
            return None
        olcu_hedef = int(float(str(olcu_sayisi)))
        olcu_idx   = None
        for i, val in enumerate(olcu_row):
            try:
                if val is not None and int(float(str(val))) == olcu_hedef:
                    olcu_idx = i
                    break
            except Exception:
                pass
        if olcu_idx is None:
            return None
        en    = int(float(str(model_satiri[olcu_idx - 3])))
        boy   = int(float(str(model_satiri[olcu_idx - 2])))
        yuk   = int(float(str(model_satiri[olcu_idx - 1])))
        hacim = round(en * boy * yuk / 1_000_000, 4)
        return {"En": en, "Boy": boy, "Yukseklik": yuk, "Tekil_Hacim": hacim}
    except Exception:
        return None


def parcalari_olustur(row, parca_haritasi, siparis_adedi):
    parcalar = []
    for parca in parca_haritasi:
        if parca["Adet_Idx"] >= len(row):
            continue
        model = row.iloc[parca["Model_Idx"]]
        olcu  = row.iloc[parca["Olcu_Idx"]]
        adet  = row.iloc[parca["Adet_Idx"]]
        if pd.notna(adet) and pd.notna(model) and str(model).strip() not in ["-", "", "nan"]:
            try:
                adet_s = int(float(str(adet).strip()))
                olcu_s = int(float(str(olcu).strip()))
                if adet_s > 0:
                    toplam_adet   = adet_s * siparis_adedi
                    boyutlar      = cbm_tablosundan_boyut_cek(parca["Kategori"], model, olcu_s)
                    if boyutlar:
                        birim_hacim   = boyutlar["Tekil_Hacim"]
                        toplam_hacim  = round(birim_hacim * toplam_adet, 4)
                        paketli_hacim = round(toplam_hacim * PAKET_KATSAYI, 4)
                        parcalar.append({
                            "kategori":      parca["Kategori"],
                            "ad":            parca["Kategori"] + " (" + str(model).strip() + ")",
                            "model":         str(model).strip(),
                            "adet":          toplam_adet,
                            "birim_hacim":   birim_hacim,
                            "toplam_hacim":  toplam_hacim,
                            "paketli_hacim": paketli_hacim,
                            "yukseklik":     boyutlar["Yukseklik"],
                            "en":            boyutlar["En"],
                            "boy":           boyutlar["Boy"],
                        })
            except Exception:
                pass
    return parcalar


PARCA_HARITASI = [
    {"Kategori": "BAZA",   "Model_Idx": 3,  "Olcu_Idx": 4,  "Adet_Idx": 5},
    {"Kategori": "BASLIK", "Model_Idx": 6,  "Olcu_Idx": 7,  "Adet_Idx": 8},
    {"Kategori": "YATAK",  "Model_Idx": 9,  "Olcu_Idx": 10, "Adet_Idx": 11},
    {"Kategori": "TOPPER", "Model_Idx": 12, "Olcu_Idx": 13, "Adet_Idx": 14},
]


def baza_tipi_onceligi(model_adi):
    ad = model_adi.upper()
    if "KAPALI" in ad or "YAYLI" in ad:
        return 0
    return 1


def en_iyi_oryantasyon(p_en, p_boy, arac_en):
    a = int(arac_en / p_en) if p_en else 0
    b = int(arac_en / p_boy) if p_boy else 0
    if b > a:
        return b, p_en
    return a, p_boy


def yukleme_plani_hesapla(set_gruplari, arac_en, arac_boy, arac_yukseklik):
    """
    Her dilimde:
      DİK KAT : bazalar/baslıklar dik → kalınlık (yukseklik) arac enine
      DÜZ KAT : ust kalan yukseklige aynı bazalar duz → daha az dilim
    """
    kuyruk_ham       = []
    ust_kat_parcalar = []
    max_baza_boy     = 200

    for grup in set_gruplari:
        boyut = grup["boyut"]
        for p in grup["parcalar"]:
            kat = p["kategori"]
            if kat in ("BAZA", "BASLIK"):
                oncelik = baza_tipi_onceligi(p["model"]) if kat == "BAZA" else 2
                kuyruk_ham.append({
                    "ad":        p["ad"],
                    "kategori":  kat,
                    "model":     p["model"],
                    "boyut":     boyut,
                    "adet":      p["adet"],
                    "yukseklik": p["yukseklik"],
                    "baza_en":   p["en"],
                    "baza_boy":  p["boy"],
                    "oncelik":   oncelik,
                })
                if kat == "BAZA":
                    max_baza_boy = max(max_baza_boy, p["boy"])
            elif kat in ("YATAK", "TOPPER"):
                ust_kat_parcalar.append({
                    "ad": p["ad"], "kategori": kat, "boyut": boyut,
                    "adet": p["adet"], "en": p["en"],
                    "boy": p["boy"], "yukseklik": p["yukseklik"],
                })

    kalan_yukseklik = max(0, arac_yukseklik - max_baza_boy)

    kuyruk = sorted(kuyruk_ham, key=lambda x: (x["oncelik"], x["yukseklik"], x["boyut"]))
    for item in kuyruk:
        item["kalan"] = item["adet"]

    dilimler        = []
    dilim_no        = 0
    konum           = 0
    dil_dik         = []
    dil_duz         = []
    dil_dik_kalan   = arac_en
    dil_duz_h_kalan = kalan_yukseklik

    def duz_kapasitesi(item):
        return int(arac_en / item["baza_boy"]) if item.get("baza_boy") else 0

    def dilimi_kapat():
        nonlocal dilim_no, konum, dil_dik_kalan, dil_duz_h_kalan
        if not dil_dik and not dil_duz:
            return
        tum_ic   = dil_dik + dil_duz
        derinlik = max(ic["baza_en"] for ic in tum_ic)
        toplam   = sum(ic["adet"] for ic in tum_ic)
        kul_en   = sum(ic["adet"] * ic["yukseklik"] for ic in dil_dik)
        dilim_no += 1
        dilimler.append({
            "dilim_no":        dilim_no,
            "konum_baslangic": konum,
            "konum_bitis":     konum + derinlik,
            "derinlik":        derinlik,
            "kullanilan_en":   kul_en,
            "arac_en":         arac_en,
            "toplam_sira":     toplam,
            "icerik":          list(dil_dik),
            "duz_icerik":      list(dil_duz),
        })
        konum           += derinlik
        dil_dik_kalan    = arac_en
        dil_duz_h_kalan  = kalan_yukseklik

    def ekle_ic(liste, item, adet, duz=False):
        for ic in liste:
            if ic["ad"] == item["ad"] and ic["boyut"] == item["boyut"]:
                ic["adet"] += adet
                return
        d = {
            "ad": item["ad"], "kategori": item["kategori"],
            "boyut": item["boyut"], "adet": adet,
            "yukseklik": item["yukseklik"], "baza_en": item["baza_en"],
        }
        if duz:
            d["duz"] = True
        liste.append(d)

    kamyon_dolu = False
    tasmalar    = []

    for item in kuyruk:
        if kamyon_dolu:
            if item["kalan"] > 0:
                tasmalar.append(dict(item))
            continue

        while item["kalan"] > 0:
            # 1. DIK KAT
            dik_slots = int(dil_dik_kalan / item["yukseklik"])
            if dik_slots > 0:
                konan = min(item["kalan"], dik_slots)
                ekle_ic(dil_dik, item, konan)
                dil_dik_kalan -= konan * item["yukseklik"]
                item["kalan"] -= konan
                continue

            # 2. DUZ KAT (ust yukseklik)
            duz_pk  = duz_kapasitesi(item)
            duz_kat = int(dil_duz_h_kalan / item["yukseklik"]) if item["yukseklik"] else 0
            duz_slots = duz_pk * duz_kat
            if duz_slots > 0:
                konan   = min(item["kalan"], duz_slots)
                k_kullan = math.ceil(konan / duz_pk) if duz_pk else 0
                ekle_ic(dil_duz, item, konan, duz=True)
                dil_duz_h_kalan -= k_kullan * item["yukseklik"]
                item["kalan"]   -= konan
                continue

            # 3. Dilim dolu → kapat, yenisini ac
            sonraki_der = item["baza_en"]
            if dil_dik:
                sonraki_der = max(ic["baza_en"] for ic in dil_dik)
            if konum + sonraki_der > arac_boy:
                kamyon_dolu = True
                tasmalar.append(dict(item))
                break
            dilimi_kapat()
            dil_dik.clear()
            dil_duz.clear()
            dil_dik_kalan   = arac_en
            dil_duz_h_kalan = kalan_yukseklik

        if kamyon_dolu and item.get("kalan", 0) > 0:
            tasmalar.append(dict(item))

    if dil_dik or dil_duz:
        dilimi_kapat()
        dil_dik.clear()
        dil_duz.clear()

    # Ust kat: tasmalar once, sonra yatak/topper
    ust_kat_hesap = []
    for t in tasmalar:
        duz_pk = duz_kapasitesi(t) if "baza_boy" in t else 1
        katlar = int(kalan_yukseklik / t["yukseklik"]) if t.get("yukseklik") else 0
        ust_kat_hesap.append({
            "tip": "BAZA_FLAT", "ad": t["ad"], "kategori": t.get("kategori", "BAZA"),
            "boyut": t["boyut"], "adet": t.get("kalan", t.get("adet", 0)),
            "yukseklik": t["yukseklik"], "kalan_yukseklik": kalan_yukseklik,
            "yan_yana": duz_pk, "kat": katlar,
        })
    for p in ust_kat_parcalar:
        yan_yana, _ = en_iyi_oryantasyon(p["en"], p["boy"], arac_en)
        kat = int(kalan_yukseklik / p["yukseklik"]) if p["yukseklik"] and kalan_yukseklik > 0 else 0
        ust_kat_hesap.append({
            "tip": p["kategori"], "ad": p["ad"], "kategori": p["kategori"],
            "boyut": p["boyut"], "adet": p["adet"],
            "en": p["en"], "boy": p["boy"], "yukseklik": p["yukseklik"],
            "kalan_yukseklik": kalan_yukseklik, "yan_yana": yan_yana, "kat": kat,
        })

    return {
        "dilimler":        dilimler,
        "ust_kat":         ust_kat_hesap,
        "kalan_yukseklik": kalan_yukseklik,
    }


class YuklemePlaniPenceresi(QDialog):
    def __init__(self, set_gruplari, parent=None):
        super().__init__(parent)
        self.set_gruplari = set_gruplari
        self.setWindowTitle("Yukleme Plani")
        self.setMinimumSize(1100, 650)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        arac_kutu   = QGroupBox("Arac Tipi")
        arac_layout = QHBoxLayout(arac_kutu)
        self.btn_group     = QButtonGroup(self)
        self.radio_buttons = {}
        for i, arac_adi in enumerate(ARAC_BOYUTLARI.keys()):
            rb = QRadioButton(arac_adi)
            if i == 0:
                rb.setChecked(True)
            self.btn_group.addButton(rb)
            self.radio_buttons[arac_adi] = rb
            arac_layout.addWidget(rb)
        layout.addWidget(arac_kutu)

        self.btn_plan = QPushButton("Yukleme Planini Olustur")
        self.btn_plan.setStyleSheet(
            "background-color:#1a5276;color:white;font-weight:bold;padding:8px;font-size:13px;")
        self.btn_plan.clicked.connect(self.plani_olustur)
        layout.addWidget(self.btn_plan)

        self.lbl_ozet = QLabel("")
        self.lbl_ozet.setStyleSheet("font-size:12px;font-weight:bold;color:#1a5276;")
        layout.addWidget(self.lbl_ozet)

        self.tablo = QTableWidget()
        self.tablo.setColumnCount(6)
        self.tablo.setHorizontalHeaderLabels(
            ["SIRA", "KONUM / DER.", "ICERIK", "SIRALANAN", "KULLANILAN EN (DIK)", "DOLULUK"])
        hh = self.tablo.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        hh.setStyleSheet(
            "QHeaderView::section{background-color:#1a5276;color:white;"
            "font-weight:bold;padding:5px;border:1px solid #154360;}")
        self.tablo.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tablo.verticalHeader().setVisible(False)
        layout.addWidget(self.tablo)

        kapat = QPushButton("Kapat")
        kapat.clicked.connect(self.close)
        layout.addWidget(kapat, alignment=Qt.AlignmentFlag.AlignRight)

    def secilen_arac(self):
        for adi, rb in self.radio_buttons.items():
            if rb.isChecked():
                return adi, ARAC_BOYUTLARI[adi]
        return None, None

    def plani_olustur(self):
        arac_adi, arac = self.secilen_arac()
        if not arac:
            return

        sonuc           = yukleme_plani_hesapla(
            self.set_gruplari, arac["en"], arac["boy"], arac["yukseklik"])
        dilimler        = sonuc["dilimler"]
        ust_kat         = sonuc["ust_kat"]
        kalan_yukseklik = sonuc["kalan_yukseklik"]

        if not dilimler and not ust_kat:
            self.lbl_ozet.setText("Listede planlanacak urun bulunamadi.")
            self.tablo.setRowCount(0)
            return

        toplam_uzunluk = sum(d["derinlik"] for d in dilimler)
        kalan_boy      = arac["boy"] - toplam_uzunluk
        toplam_adet    = sum(d["toplam_sira"] for d in dilimler)
        ozet = (arac_adi + "  |  " + str(len(dilimler)) + " dilim  |  " +
                str(toplam_adet) + " adet  |  Kullanilan boy: " +
                str(toplam_uzunluk) + " cm  |  Kalan: " + str(kalan_boy) + " cm")

        RENKLER     = [QColor("#d6eaf8"), QColor("#d5f5e3"), QColor("#fdebd0"),
                       QColor("#f9ebea"), QColor("#e8daef"), QColor("#d0ece7")]
        DUZ_TINT    = QColor("#fffde7")
        UST_KAT_BG  = QColor("#1d6a3a")
        UST_KAT_FG  = QColor("#ffffff")
        bold = QFont()
        bold.setBold(True)
        italic = QFont()
        italic.setItalic(True)

        self.tablo.setRowCount(0)
        renk_idx = 0
        son_tip  = None

        for d in dilimler:
            r = self.tablo.rowCount()
            self.tablo.insertRow(r)

            ilk_tip = d["icerik"][0]["ad"] if d["icerik"] else (
                d["duz_icerik"][0]["ad"] if d.get("duz_icerik") else "")
            if ilk_tip != son_tip:
                renk_idx += 1
                son_tip = ilk_tip
            bg = RENKLER[renk_idx % len(RENKLER)]

            # Dik kat icerigi
            dik_parcalar = []
            for ic in d["icerik"]:
                dik_parcalar.append(
                    ic["ad"] + " " + ic["boyut"] +
                    " x" + str(ic["adet"]) + " (" + str(ic["yukseklik"]) + "cm)")
            # Duz kat icerigi
            duz_parcalar = []
            for ic in d.get("duz_icerik", []):
                duz_parcalar.append(
                    "[DUZ] " + ic["ad"] + " " + ic["boyut"] +
                    " x" + str(ic["adet"]) + " (" + str(ic["yukseklik"]) + "cm)")
            icerik_metni = "  |  ".join(dik_parcalar + duz_parcalar)

            doluluk_pct    = round(d["kullanilan_en"] / d["arac_en"] * 100)
            kullanilan_str = str(d["kullanilan_en"]) + " / " + str(d["arac_en"]) + " cm"
            konum_metni    = (str(d["konum_baslangic"]) + "-" + str(d["konum_bitis"]) +
                              " (" + str(d["derinlik"]) + "cm)")

            has_duz = bool(d.get("duz_icerik"))
            row_bg  = DUZ_TINT if has_duz else bg

            def hucre(metin, sag=False, kalin=False, _bg=row_bg):
                item = QTableWidgetItem(str(metin))
                item.setBackground(_bg)
                item.setTextAlignment(
                    (Qt.AlignmentFlag.AlignRight if sag else Qt.AlignmentFlag.AlignCenter)
                    | Qt.AlignmentFlag.AlignVCenter)
                if kalin:
                    item.setFont(bold)
                return item

            self.tablo.setItem(r, 0, hucre(d["dilim_no"], kalin=True))
            self.tablo.setItem(r, 1, hucre(konum_metni))
            self.tablo.setItem(r, 2, hucre(icerik_metni))
            self.tablo.setItem(r, 3, hucre(d["toplam_sira"]))
            self.tablo.setItem(r, 4, hucre(kullanilan_str))
            self.tablo.setItem(r, 5, hucre("%" + str(doluluk_pct), kalin=(doluluk_pct < 90)))
            self.tablo.setRowHeight(r, 28)

        # Ust kat bolumu
        if ust_kat:
            r = self.tablo.rowCount()
            self.tablo.insertRow(r)
            self.tablo.setSpan(r, 0, 1, 6)
            bi = QTableWidgetItem(
                "  UST KAT  —  Baza ustunde kalan yukseklik: " +
                str(kalan_yukseklik) + " cm  |  " +
                ("Once baza/baslik tasma, sonra yatak/topper" if any(
                    u["tip"] == "BAZA_FLAT" for u in ust_kat)
                 else "Baza kalmadi, yatak / topper"))
            bi.setBackground(UST_KAT_BG)
            bi.setForeground(UST_KAT_FG)
            bi.setFont(bold)
            self.tablo.setItem(r, 0, bi)
            self.tablo.setRowHeight(r, 28)

            for p in ust_kat:
                r = self.tablo.rowCount()
                self.tablo.insertRow(r)
                bg2 = (QColor("#fdebd0") if p["tip"] == "BAZA_FLAT" else QColor("#d0ece7"))

                def hucre2(metin, sag=False, kalin=False, _bg=bg2):
                    item = QTableWidgetItem(str(metin))
                    item.setBackground(_bg)
                    item.setTextAlignment(
                        (Qt.AlignmentFlag.AlignRight if sag else Qt.AlignmentFlag.AlignCenter)
                        | Qt.AlignmentFlag.AlignVCenter)
                    if kalin:
                        item.setFont(bold)
                    return item

                kat      = p.get("kat", 0)
                yy       = p.get("yan_yana", 1)
                boyutlar = (str(p["yukseklik"]) + "cm kalin (duz)" if p["tip"] == "BAZA_FLAT"
                            else str(p.get("en","")) + "x" + str(p.get("boy","")) +
                                 "x" + str(p["yukseklik"]) + " cm")
                siralama = (str(yy) + " yan yana x " + str(kat) + " kat = " +
                            str(yy * kat) + " adet/poz.")

                self.tablo.setItem(r, 0, hucre2(p["tip"], kalin=True))
                self.tablo.setItem(r, 1, hucre2(p["boyut"]))
                self.tablo.setItem(r, 2, hucre2(p["ad"] + "  (" + boyutlar + ")"))
                self.tablo.setItem(r, 3, hucre2(str(p["adet"]) + " adet"))
                self.tablo.setItem(r, 4, hucre2(siralama))
                self.tablo.setItem(r, 5, hucre2(""))
                self.tablo.setRowHeight(r, 26)

        if kalan_boy < 0:
            self.lbl_ozet.setStyleSheet("font-size:12px;font-weight:bold;color:#c0392b;")
            self.lbl_ozet.setText(ozet + "  ARAC YETMIYOR!")
        else:
            self.lbl_ozet.setStyleSheet("font-size:12px;font-weight:bold;color:#1a5276;")
            self.lbl_ozet.setText(ozet)


class TirYuklemeUygulamasi(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("CBM Programi")
        self.setGeometry(100, 100, 1100, 700)
        self.matris_df           = pd.read_excel(TANIM_EXCEL, header=0)
        self.set_gruplari        = []
        self.satir_grup_haritasi = {}
        self.init_ui()
        self.firma_listesini_guncelle()

    def init_ui(self):
        ana_widget = QWidget()
        self.setCentralWidget(ana_widget)
        ana_layout = QVBoxLayout(ana_widget)
        ana_layout.setContentsMargins(12, 12, 12, 12)
        ana_layout.setSpacing(8)

        secim_layout = QHBoxLayout()
        secim_layout.setSpacing(12)
        for etiket, attr in [
            ("1. Firma Secimi:",  "combo_firma"),
            ("2. Set Secimi:",    "combo_set"),
            ("3. Boyut Secimi:",  "combo_boyut"),
        ]:
            kutu = QVBoxLayout()
            kutu.addWidget(QLabel(etiket))
            combo = QComboBox()
            setattr(self, attr, combo)
            kutu.addWidget(combo)
            secim_layout.addLayout(kutu)

        self.combo_firma.currentTextChanged.connect(self.set_listesini_guncelle)
        self.combo_firma.currentTextChanged.connect(lambda _: self._baslik_guncelle())
        self.combo_set.currentTextChanged.connect(self.boyut_listesini_guncelle)

        adet_kutu = QVBoxLayout()
        adet_kutu.addWidget(QLabel("4. Siparis Adedi:"))
        self.spin_adet = QSpinBox()
        self.spin_adet.setMinimum(1)
        self.spin_adet.setMaximum(9999)
        adet_kutu.addWidget(self.spin_adet)
        secim_layout.addLayout(adet_kutu)

        btn_kutu = QVBoxLayout()
        btn_kutu.addWidget(QLabel(""))
        self.btn_ekle = QPushButton("Listeye Ekle")
        self.btn_ekle.setStyleSheet(
            "background-color:#2ecc71;color:white;font-weight:bold;padding:6px 16px;")
        self.btn_ekle.clicked.connect(self.set_hesapla_ve_ekle)
        btn_kutu.addWidget(self.btn_ekle)
        self.btn_temizle = QPushButton("Listeyi Temizle")
        self.btn_temizle.setStyleSheet(
            "background-color:#e74c3c;color:white;font-weight:bold;padding:6px 16px;")
        self.btn_temizle.clicked.connect(self.listeyi_temizle)
        btn_kutu.addWidget(self.btn_temizle)
        secim_layout.addLayout(btn_kutu)

        plan_kutu = QVBoxLayout()
        plan_kutu.addWidget(QLabel(""))
        self.btn_plan = QPushButton("Yukleme Plani")
        self.btn_plan.setStyleSheet(
            "background-color:#1a5276;color:white;font-weight:bold;padding:6px 16px;")
        self.btn_plan.clicked.connect(self.yukleme_planini_ac)
        plan_kutu.addWidget(self.btn_plan)
        plan_kutu.addWidget(QLabel(""))
        secim_layout.addLayout(plan_kutu)
        ana_layout.addLayout(secim_layout)

        isim_layout = QHBoxLayout()
        isim_layout.addWidget(QLabel("Tablo Basligi:"))
        self.edit_tablo_ismi = QLineEdit()
        self.edit_tablo_ismi.setPlaceholderText("orn: KEY WEST 20.05.2026 TUM SIPARISLERI")
        self.edit_tablo_ismi.setStyleSheet("font-size:12px;padding:4px;")
        self.edit_tablo_ismi.textChanged.connect(lambda _: self._baslik_guncelle())
        isim_layout.addWidget(self.edit_tablo_ismi)
        ana_layout.addLayout(isim_layout)

        self.lbl_baslik = QLabel()
        self.lbl_baslik.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_baslik.setStyleSheet(
            "font-size:16px;font-weight:bold;color:white;"
            "background-color:#1a5276;padding:8px;border-radius:4px;")
        ana_layout.addWidget(self.lbl_baslik)
        self._baslik_guncelle()

        self.tablo = QTableWidget()
        self.tablo.setColumnCount(6)
        self.tablo.setHorizontalHeaderLabels(
            ["SIRA", "URUN ADI", "ADET",
             "BIRIM HACIM (m3)", "TOPLAM HACIM (m3)", "PAKETLI HACIM (m3)"])
        hh = self.tablo.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        hh.setStyleSheet(
            "QHeaderView::section{background-color:#1a5276;color:white;"
            "font-weight:bold;padding:5px;border:1px solid #154360;}")
        self.tablo.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tablo.setAlternatingRowColors(False)
        self.tablo.verticalHeader().setVisible(False)
        self.tablo.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tablo.customContextMenuRequested.connect(self.sag_tik_menusu)
        ana_layout.addWidget(self.tablo)

        self.lbl_toplam = QLabel(
            "TOPLAM HACIM: 0.0000 m3     PAKETLI TOPLAM: 0.0000 m3")
        self.lbl_toplam.setStyleSheet(
            "font-size:13px;font-weight:bold;color:#1a5276;margin-top:6px;")
        ana_layout.addWidget(self.lbl_toplam, alignment=Qt.AlignmentFlag.AlignRight)

    def yukleme_planini_ac(self):
        if not self.set_gruplari:
            QMessageBox.information(self, "Bilgi", "Once listeye urun ekleyin.")
            return
        pencere = YuklemePlaniPenceresi(self.set_gruplari, self)
        pencere.exec()

    def _baslik_guncelle(self):
        ozel = self.edit_tablo_ismi.text().strip() if hasattr(self, "edit_tablo_ismi") else ""
        if ozel:
            self.lbl_baslik.setText(ozel)
        else:
            firma = self.combo_firma.currentText() if self.combo_firma.count() > 0 else ""
            tarih = QDate.currentDate().toString("dd.MM.yyyy")
            self.lbl_baslik.setText(firma + "  " + tarih + "  CBM TABLOSU")

    def firma_listesini_guncelle(self):
        self.combo_firma.blockSignals(True)
        firmalar = self.matris_df.iloc[:, 0].dropna().unique().tolist()
        self.combo_firma.clear()
        self.combo_firma.addItems([str(f).strip() for f in firmalar if str(f).strip()])
        self.combo_firma.blockSignals(False)
        if firmalar:
            self.set_listesini_guncelle(self.combo_firma.currentText())
        self._baslik_guncelle()

    def set_listesini_guncelle(self, secilen_firma):
        self.combo_set.blockSignals(True)
        filtre = (self.matris_df.iloc[:, 0].astype(str)
                  .str.replace(" ", "").str.upper()
                  == str(secilen_firma).replace(" ", "").upper())
        setler = self.matris_df[filtre].iloc[:, 1].dropna().unique().tolist()
        self.combo_set.clear()
        self.combo_set.addItems([str(s).strip() for s in setler])
        self.combo_set.blockSignals(False)
        if setler:
            self.boyut_listesini_guncelle(self.combo_set.currentText())

    def boyut_listesini_guncelle(self, secilen_set):
        secilen_firma = self.combo_firma.currentText()
        filtre = (
            (self.matris_df.iloc[:, 0].astype(str).str.replace(" ", "").str.upper()
             == str(secilen_firma).replace(" ", "").upper()) &
            (self.matris_df.iloc[:, 1].astype(str).str.replace(" ", "").str.upper()
             == str(secilen_set).replace(" ", "").upper())
        )
        boyutlar = self.matris_df[filtre].iloc[:, 2].dropna().unique().tolist()
        self.combo_boyut.clear()
        self.combo_boyut.addItems([str(b).strip() for b in boyutlar])

    def sag_tik_menusu(self, pos):
        satir = self.tablo.rowAt(pos.y())
        if satir < 0:
            return
        grup_idx = self.satir_grup_haritasi.get(satir)
        if grup_idx is None:
            return
        menu = QMenu(self)
        eylem_duzenle = menu.addAction("Adedi Duzenle")
        eylem_sil     = menu.addAction("Grubu Sil")
        eylem_plan    = menu.addAction("Yukleme Planini Goster")
        secilen = menu.exec(self.tablo.viewport().mapToGlobal(pos))
        if secilen == eylem_duzenle:
            grup = self.set_gruplari[grup_idx]
            yeni_adet, tamam = QInputDialog.getInt(
                self, "Adedi Duzenle",
                grup["firma"] + " - " + grup["set_adi"] + " " + grup["boyut"] +
                "\nYeni siparis adedi:",
                value=grup["siparis_adedi"], min=1, max=9999)
            if tamam and yeni_adet != grup["siparis_adedi"]:
                self._grubu_yeniden_hesapla(grup_idx, yeni_adet)
                self.tabloyu_yeniden_ciz()
        elif secilen == eylem_sil:
            del self.set_gruplari[grup_idx]
            self.tabloyu_yeniden_ciz()
        elif secilen == eylem_plan:
            self.yukleme_planini_ac()

    def _grubu_yeniden_hesapla(self, grup_idx, yeni_adet):
        grup = self.set_gruplari[grup_idx]
        filtre = (
            (self.matris_df.iloc[:, 0].astype(str).str.replace(" ", "").str.upper()
             == str(grup["firma"]).replace(" ", "").upper()) &
            (self.matris_df.iloc[:, 1].astype(str).str.replace(" ", "").str.upper()
             == str(grup["set_adi"]).replace(" ", "").upper()) &
            (self.matris_df.iloc[:, 2].astype(str).str.replace(" ", "").str.upper()
             == str(grup["boyut"]).replace(" ", "").upper())
        )
        secilen_satir = self.matris_df[filtre]
        if secilen_satir.empty:
            return
        row = secilen_satir.iloc[0]
        grup["siparis_adedi"] = yeni_adet
        grup["parcalar"]      = parcalari_olustur(row, PARCA_HARITASI, yeni_adet)

    def set_hesapla_ve_ekle(self):
        firma         = self.combo_firma.currentText()
        set_adi       = self.combo_set.currentText()
        boyut         = self.combo_boyut.currentText()
        siparis_adedi = self.spin_adet.value()
        filtre = (
            (self.matris_df.iloc[:, 0].astype(str).str.replace(" ", "").str.upper()
             == str(firma).replace(" ", "").upper()) &
            (self.matris_df.iloc[:, 1].astype(str).str.replace(" ", "").str.upper()
             == str(set_adi).replace(" ", "").upper()) &
            (self.matris_df.iloc[:, 2].astype(str).str.replace(" ", "").str.upper()
             == str(boyut).replace(" ", "").upper())
        )
        secilen_satir = self.matris_df[filtre]
        if secilen_satir.empty:
            QMessageBox.warning(self, "Hata", "Secilen kriterlere uygun set bulunamadi.")
            return
        row      = secilen_satir.iloc[0]
        parcalar = parcalari_olustur(row, PARCA_HARITASI, siparis_adedi)
        if parcalar:
            self.set_gruplari.append({
                "firma":         firma,
                "set_adi":       set_adi,
                "boyut":         boyut,
                "siparis_adedi": siparis_adedi,
                "parcalar":      parcalar,
            })
            self.tabloyu_yeniden_ciz()
        else:
            QMessageBox.information(self, "Bilgi", "Bu set icin CBM karsiligi uretilemedi.")

    def listeyi_temizle(self):
        self.set_gruplari.clear()
        self.tabloyu_yeniden_ciz()

    def tabloyu_yeniden_ciz(self):
        self.tablo.setRowCount(0)
        self.satir_grup_haritasi = {}

        BASLIK_BG  = QColor("#1a5276")
        BASLIK_FG  = QColor("#ffffff")
        PARCA_BG_1 = QColor("#d6eaf8")
        PARCA_BG_2 = QColor("#eaf4fb")
        TOPLAM_BG  = QColor("#154360")
        TOPLAM_FG  = QColor("#ffffff")

        bold = QFont()
        bold.setBold(True)

        toplam_genel  = 0.0
        paketli_genel = 0.0
        sira_no       = 0

        for grup_idx, grup in enumerate(self.set_gruplari):
            parca_ozeti  = "  /  " + "  -  ".join(p["ad"] for p in grup["parcalar"])
            baslik_metni = ("  " + grup["firma"] + "  -  " + grup["set_adi"] +
                            "  " + grup["boyut"] + "  x  " +
                            str(grup["siparis_adedi"]) + " Adet" + parca_ozeti)

            r = self.tablo.rowCount()
            self.tablo.insertRow(r)
            self.satir_grup_haritasi[r] = grup_idx
            self.tablo.setSpan(r, 0, 1, 6)
            bi = QTableWidgetItem(baslik_metni)
            bi.setBackground(BASLIK_BG)
            bi.setForeground(BASLIK_FG)
            bi.setFont(bold)
            self.tablo.setItem(r, 0, bi)
            self.tablo.setRowHeight(r, 28)

            grup_toplam  = 0.0
            grup_paketli = 0.0

            for idx, p in enumerate(grup["parcalar"]):
                sira_no += 1
                r = self.tablo.rowCount()
                self.tablo.insertRow(r)
                self.satir_grup_haritasi[r] = grup_idx
                bg = PARCA_BG_1 if idx % 2 == 0 else PARCA_BG_2

                def hucre(metin, sag=False, _bg=bg):
                    item = QTableWidgetItem(str(metin))
                    item.setBackground(_bg)
                    item.setTextAlignment(
                        (Qt.AlignmentFlag.AlignRight if sag else Qt.AlignmentFlag.AlignCenter)
                        | Qt.AlignmentFlag.AlignVCenter)
                    return item

                self.tablo.setItem(r, 0, hucre(sira_no))
                self.tablo.setItem(r, 1, hucre(p["ad"]))
                self.tablo.setItem(r, 2, hucre(p["adet"]))
                self.tablo.setItem(r, 3, hucre("{:.4f}".format(p["birim_hacim"]), sag=True))
                self.tablo.setItem(r, 4, hucre("{:.4f}".format(p["toplam_hacim"]), sag=True))
                self.tablo.setItem(r, 5, hucre("{:.4f}".format(p["paketli_hacim"]), sag=True))
                grup_toplam  += p["toplam_hacim"]
                grup_paketli += p["paketli_hacim"]

            grup_toplam  = round(grup_toplam, 4)
            grup_paketli = round(grup_paketli, 4)

            r = self.tablo.rowCount()
            self.tablo.insertRow(r)
            self.satir_grup_haritasi[r] = grup_idx

            def th(metin):
                item = QTableWidgetItem(str(metin))
                item.setBackground(TOPLAM_BG)
                item.setForeground(TOPLAM_FG)
                item.setFont(bold)
                item.setTextAlignment(
                    Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                return item

            self.tablo.setItem(r, 0, th(""))
            self.tablo.setItem(r, 1, th("SET TOPLAMI"))
            self.tablo.setItem(r, 2, th(""))
            self.tablo.setItem(r, 3, th(""))
            self.tablo.setItem(r, 4, th("{:.4f}".format(grup_toplam)))
            self.tablo.setItem(r, 5, th("{:.4f}".format(grup_paketli)))
            self.tablo.setRowHeight(r, 26)

            toplam_genel  += grup_toplam
            paketli_genel += grup_paketli

        # Ayak hesaplama
        toplam_ayak = 0
        for grup in self.set_gruplari:
            try:
                genislik = int(str(grup["boyut"]).split("X")[0].strip())
            except Exception:
                continue
            ayak_per_baza = 6 if genislik in AYAK_KUCUK_GENISLIKLER else 12
            for p in grup["parcalar"]:
                ad_upper = p["ad"].upper()
                if "BAZA" in ad_upper and ("YAYLI" in ad_upper or "KAPALI" in ad_upper):
                    toplam_ayak += p["adet"] * ayak_per_baza

        if toplam_ayak > 0:
            koli_sayisi = math.ceil(toplam_ayak / AYAK_KOLI_KAPASITE)
            ayak_hacim  = round(koli_sayisi * AYAK_KOLI_HACIM, 4)
            AYAK_BG = QColor("#1d6a3a")
            AYAK_FG = QColor("#ffffff")

            def ayak_hucre(metin, sag=False):
                item = QTableWidgetItem(str(metin))
                item.setBackground(AYAK_BG)
                item.setForeground(AYAK_FG)
                f = QFont()
                f.setBold(True)
                item.setFont(f)
                item.setTextAlignment(
                    (Qt.AlignmentFlag.AlignRight if sag else Qt.AlignmentFlag.AlignCenter)
                    | Qt.AlignmentFlag.AlignVCenter)
                return item

            r = self.tablo.rowCount()
            self.tablo.insertRow(r)
            self.tablo.setItem(r, 0, ayak_hucre(""))
            self.tablo.setItem(r, 1, ayak_hucre(
                "AYAK KOLISI  (" + str(toplam_ayak) + " ayak - " +
                str(koli_sayisi) + " koli x 144 luk)"))
            self.tablo.setItem(r, 2, ayak_hucre(koli_sayisi))
            self.tablo.setItem(r, 3, ayak_hucre("{:.4f}".format(AYAK_KOLI_HACIM), sag=True))
            self.tablo.setItem(r, 4, ayak_hucre("{:.4f}".format(ayak_hacim), sag=True))
            self.tablo.setItem(r, 5, ayak_hucre("{:.4f}".format(ayak_hacim), sag=True))
            self.tablo.setRowHeight(r, 28)
            toplam_genel  += ayak_hacim
            paketli_genel += ayak_hacim

        self.lbl_toplam.setText(
            "GENEL TOPLAM:  {:.4f} m3     PAKETLI TOPLAM:  {:.4f} m3".format(
                toplam_genel, round(paketli_genel, 4)))


if __name__ == "__main__":
    app = QApplication(sys.argv)
    pencere = TirYuklemeUygulamasi()
    pencere.show()
    sys.exit(app.exec())
