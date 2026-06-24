import pandas as pd
from openpyxl import load_workbook

cbm_excel = "CBM.xlsx"
tanim_excel = "set_tanimlari.xlsx"


def cbm_tablosundan_boyut_cek(sayfa_adi, model_adi, olcu_sayisi):
    """
    CBM.xlsx dosyasındaki ilgili sayfadan kutu boyutlarını ve hacmi döndürür.

    Excel yapısı (her sayfa aynı):
      Satır 0: None, 'Baza', None, None, 70, 'Baza', None, None, 80, ...
               Ölçü sayısı (70, 80, 90...) o grubun HACİM sütununun index'idir.
      Satır 1: MODEL, EN, BOY, YÜKSEKLİK, HACİM(formül), EN, BOY, ...
      Satır 2+: Veriler (HACİM sütunları formül içerdiğinden Python'da hesaplanır)

    Offset mantığı:
      olcu_idx = 0. satırda ölçü sayısının index'i (= HACİM pozisyonu)
      EN  = olcu_idx - 3
      BOY = olcu_idx - 2
      YÜK = olcu_idx - 1
      HACİM = EN * BOY * YÜK / 1_000_000  (mm³ → m³)
    """
    try:
        wb = load_workbook(cbm_excel, read_only=True, data_only=True)
        ws = wb[sayfa_adi.upper()]
        rows = list(ws.iter_rows(values_only=True))

        olcu_row = rows[0]

        # Model satırını bul
        model_aranan = str(model_adi).replace(" ", "").upper()
        model_satiri = None
        for r in rows[2:]:
            if r[0] and str(r[0]).replace(" ", "").upper() == model_aranan:
                model_satiri = r
                break

        if model_satiri is None:
            print(f"   [İpucu] {sayfa_adi.upper()} -> '{model_adi}' modeli bulunamadı!")
            return None

        # 0. satırda ölçü sayısını bul (bu index = HACİM pozisyonu)
        olcu_hedef = int(float(str(olcu_sayisi)))
        olcu_idx = None
        for i, val in enumerate(olcu_row):
            try:
                if int(float(str(val))) == olcu_hedef:
                    olcu_idx = i
                    break
            except Exception:
                pass

        if olcu_idx is None:
            print(f"   [İpucu] '{model_adi}' için {olcu_hedef} cm ölçüsü bulunamadı!")
            return None

        en  = model_satiri[olcu_idx - 3]
        boy = model_satiri[olcu_idx - 2]
        yuk = model_satiri[olcu_idx - 1]
        hacim = round(en * boy * yuk / 1_000_000, 4)

        return {"En": en, "Boy": boy, "Yükseklik": yuk, "Tekil_Hacim": hacim}

    except Exception as e:
        print(f"   [HATA] cbm_tablosundan_boyut_cek ({sayfa_adi} / {model_adi}): {e}")
        return None


def canli_set_cozucu(firma, set_adi, boyut, siparis_adedi):
    try:
        print(f"=== PROGRAM GİRİŞİ ===")
        print(f"Firma: {firma} | Set: {set_adi} | Boyut: {boyut} | Sipariş: {siparis_adedi} Adet\n")

        # set_tanimlari.xlsx: 0. satır başlıklar (FİRMA, SET ADI, BOYUT, ...)
        matris_df = pd.read_excel(tanim_excel, header=0)

        firma_temiz = str(firma).replace(" ", "").upper()
        set_temiz   = str(set_adi).replace(" ", "").upper()
        boyut_temiz = str(boyut).replace(" ", "").upper()

        filtre = (
            matris_df.iloc[:, 0].astype(str).str.replace(" ", "").str.upper() == firma_temiz
        ) & (
            matris_df.iloc[:, 1].astype(str).str.replace(" ", "").str.upper() == set_temiz
        ) & (
            matris_df.iloc[:, 2].astype(str).str.replace(" ", "").str.upper() == boyut_temiz
        )

        secilen_satir = matris_df[filtre]
        if secilen_satir.empty:
            print("Hata: Kriterlere uygun set bulunamadı!")
            return

        row = secilen_satir.iloc[0]
        yukleme_listesi = []

        # set_tanimlari sütun düzeni:
        #   0=FİRMA, 1=SET ADI, 2=BOYUT,
        #   3=BAZA MODEL,   4=BAZA ÖLÇÜ,   5=BAZA ADET,
        #   6=BAŞLIK MODEL, 7=BAŞLIK ÖLÇÜ, 8=BAŞLIK ADET,
        #   9=YATAK MODEL,  10=YATAK ÖLÇÜ, 11=YATAK ADET,
        #  12=TOPPER MODEL, 13=TOPPER ÖLÇÜ, 14=TOPPER ADET
        parca_haritasi = [
            {"Kategori": "BAZA",   "Model_Idx": 3,  "Olcu_Idx": 4,  "Adet_Idx": 5},
            {"Kategori": "BAŞLIK", "Model_Idx": 6,  "Olcu_Idx": 7,  "Adet_Idx": 8},
            {"Kategori": "YATAK",  "Model_Idx": 9,  "Olcu_Idx": 10, "Adet_Idx": 11},
            {"Kategori": "TOPPER", "Model_Idx": 12, "Olcu_Idx": 13, "Adet_Idx": 14},
        ]

        print("--- PARÇALAR ANALİZ EDİLİYOR ---")
        for parca in parca_haritasi:
            if parca["Adet_Idx"] >= len(row):
                continue

            model = row.iloc[parca["Model_Idx"]]
            olcu  = row.iloc[parca["Olcu_Idx"]]
            adet  = row.iloc[parca["Adet_Idx"]]

            if (pd.notna(adet) and pd.notna(model) and
                    str(model).strip() not in ["-", "", "nan"]):
                try:
                    adet_sayisi = int(float(str(adet).strip()))
                    olcu_sayisi = int(float(str(olcu).strip()))

                    if adet_sayisi > 0:
                        print(f"> {parca['Kategori']}: Model='{str(model).strip()}', Ölçü={olcu_sayisi}, Adet={adet_sayisi}")
                        toplam_adet = adet_sayisi * int(siparis_adedi)
                        boyutlar = cbm_tablosundan_boyut_cek(parca["Kategori"], model, olcu_sayisi)

                        if boyutlar:
                            toplam_hacim = boyutlar["Tekil_Hacim"] * toplam_adet
                            yukleme_listesi.append({
                                "Paket Türü":           f"{parca['Kategori']} ({str(model).strip()})",
                                "Ölçü":                 olcu_sayisi,
                                "Toplam Adet":          toplam_adet,
                                "Kutu (En x Boy x Yük)": f"{boyutlar['En']} x {boyutlar['Boy']} x {boyutlar['Yükseklik']}",
                                "Toplam CBM":           round(toplam_hacim, 4),
                            })
                except Exception as ex:
                    print(f"   [UYARI] {parca['Kategori']} atlandı: {ex}")

        print("\n-----------------------------------------------------")
        if yukleme_listesi:
            print("--- PARÇALAR BOYUTLANDIRILDI ---")
            sonuc_df = pd.DataFrame(yukleme_listesi)
            print(sonuc_df.to_string(index=False))
            print(f"\nTOPLAM CBM: {round(sum(x['Toplam CBM'] for x in yukleme_listesi), 4)} m³")
        else:
            print("Yüklenecek geçerli bir parça bulunamadı.")

    except Exception as e:
        import traceback
        print(f"Hata: {e}")
        traceback.print_exc()


# --- CANLI TEST ---
canli_set_cozucu(firma="KEY WEST", set_adi="BOX DELUXE", boyut="90X200", siparis_adedi=1)
